"""坐班签到表解析：从 Excel 工作表实例化教师对象"""

import re
from datetime import date, datetime, time
from typing import Any

from loguru import logger
from openpyxl.worksheet.worksheet import Worksheet

from core.allowance.check.modules import Check, Teacher, Workday

# 每位教师固定 6 行：上午签到/签退、下午签到/签退、晚上签到/签退
ROWS_PER_TEACHER = 6


def parse_time(value: Any) -> time | None:
    """解析打卡时间；'-'、NaN、None、空等视为无打卡。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = str(value).strip()
    if not text or text in ("-", "nan", "NaN", "None", "null"):
        return None
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return None
    return time(int(match.group(1)), int(match.group(2)))


def parse_date_header(header: Any) -> date | None:
    """
    从列头解析日期，如 '周一 26-03-09' -> date(2026, 3, 9)。

    星期与日期之间可以是空格或回车/换行等任意空白（实际表中常为单元格内换行）；
    日期支持 yyyy-mm-dd / yy-mm-dd（两位年份按 20xx），分隔符支持 '-'、'/'、'.'。
    openpyxl 把日期列头读成 datetime 时直接取其日期部分。
    解析失败返回 None。
    """
    if header is None:
        return None
    if isinstance(header, datetime):
        return header.date()
    if isinstance(header, date):
        return header
    text = str(header).strip()
    match = re.search(r"(\d{4}|\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def load_teachers(sheet: Worksheet) -> list[Teacher]:
    """
    从签到表实例化教师列表。

    表结构：第 1 行为表头（姓名、部门、各日期列，日期列格式如 '周一 26-03-09'）；
    每位教师固定 6 行数据：
      第 1-2 行 → 上午签到/签退
      第 3-4 行 → 下午签到/签退
      第 5-6 行 → 晚上签到/签退
    姓名/部门只出现在每组首行（合并单元格）。

    仅按行位置实例化，不做离校/返校判断（返校日数据虽位于前两行，仍按上午位解析，
    由统计阶段结合假期表再按晚上规则判定）。

    Raises:
        ValueError: 第 1 行未找到任何日期列
    """
    # 定位数据列：第 1 行中能解析出日期的列
    data_cols: list[tuple[int, date]] = []
    for col in range(1, sheet.max_column + 1):
        day = parse_date_header(sheet.cell(row=2, column=col).value)
        if day is not None:
            data_cols.append((col, day))
    if not data_cols:
        raise ValueError("签到表第 1 行未找到日期列（如 '周一 26-03-09'）")
    data_cols.sort(key=lambda item: item[1])

    teachers: list[Teacher] = []
    row = 3
    while row <= sheet.max_row:
        name = sheet.cell(row=row, column=1).value
        if not name:
            row += 1
            continue
        department = sheet.cell(row=row, column=2).value
        logger.debug(f"解析教师: {name} (部门: {department})")

        workdays: list[Workday] = []
        for col, day in data_cols:
            morning = Check(
                time_in=parse_time(sheet.cell(row=row, column=col).value),
                time_out=parse_time(sheet.cell(row=row + 1, column=col).value),
            )
            afternoon = Check(
                time_in=parse_time(sheet.cell(row=row + 2, column=col).value),
                time_out=parse_time(sheet.cell(row=row + 3, column=col).value),
            )
            evening = Check(
                time_in=parse_time(sheet.cell(row=row + 4, column=col).value),
                time_out=parse_time(sheet.cell(row=row + 5, column=col).value),
            )
            workdays.append(
                Workday(
                    date=datetime.combine(day, time.min),
                    morning=morning,
                    afternoon=afternoon,
                    evening=evening,
                )
            )

        teachers.append(
            Teacher(name=str(name), department=str(department or ""), workdays=workdays)
        )
        row += ROWS_PER_TEACHER

    if not teachers:
        raise ValueError("签到表中未解析到任何教师数据")
    return teachers
