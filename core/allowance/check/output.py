"""坐班签到统计输出：生成统计 Excel 文件"""

from datetime import date, datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook

from core.allowance.check.modules import Teacher
from core.config import OUTPUT_DIR

# 输出目录：output/津贴/坐班签到统计
OUTPUT_SUBDIR = OUTPUT_DIR / "津贴" / "坐班签到统计"


def _unique_path(directory: Path, filename: str) -> Path:
    """同名文件时在文件名后追加保存时间戳，返回唯一路径。"""
    directory.mkdir(parents=True, exist_ok=True)
    candidate = directory / f"{filename}.xlsx"
    if not candidate.exists():
        return candidate
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return directory / f"{filename}_{timestamp}.xlsx"


def output_statistics(teachers: list[Teacher], date_range: tuple[date, date]) -> list[Path]:
    """
    生成签到统计 Excel 文件。

    每人一行（姓名、签到次数）；按部门分组分别生成文件。
    文件结构：第 1 行标题 `{部门}{日期范围}签到统计`，
    第 2 行表头（姓名、签到次数），第 3 行起数据。
    保存到 OUTPUT_DIR/津贴/坐班签到统计；同名文件追加时间戳。

    :return: 生成的 Excel 文件路径列表
    """
    # 按部门分组（空部门按"未分组"处理）
    groups: dict[str, list[Teacher]] = {}
    for teacher in teachers:
        department = teacher.department or "未分组"
        groups.setdefault(department, []).append(teacher)

    paths: list[Path] = []
    for department, group in groups.items():
        title = f"{department}{date_range[0]}至{date_range[1]}签到统计"

        wb = Workbook()
        ws = wb.active
        ws.title = "签到统计"
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=2, column=1, value="姓名")
        ws.cell(row=2, column=2, value="签到次数")
        for index, teacher in enumerate(group, start=3):
            ws.cell(row=index, column=1, value=teacher.name)
            ws.cell(row=index, column=2, value=teacher.count)

        path = _unique_path(OUTPUT_SUBDIR, title)
        wb.save(path)
        logger.info(f"✅ 签到统计已保存: {path}")
        paths.append(path)

    return paths
