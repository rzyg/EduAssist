import re
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.score.models import ClassStatistics, StreamingMap, SubjectStatistics
from loguru import logger

# ── 模块级常量 ──────────────────────────────────────────────────────

SUBJECT_LIST = [
    "总分",
    "语文",
    "数学",
    "英语",
    "物理",
    "化学",
    "生物",
    "历史",
    "政治",
    "地理",
    "小语种",
]

TOTAL_LINE = ["清北线", "985线", "211线", "特控线", "本科线"]


# ===================================================================
# 内部助手函数
# ===================================================================


def _filter_subjects(direction: str) -> List[str]:
    """
    根据选科方向返回实际有效的科目列表，移除另一方向的科目。
    """
    subjects = SUBJECT_LIST[:]
    if direction == "物理":
        subjects.remove("历史")
    elif direction == "历史":
        subjects.remove("物理")
    return subjects


def _extract_class_number(name: str) -> int:
    """
    从班级名中提取所有数字并转换为整数，用于排序。

    例: "2510" → 2510, "2510班" → 2510, "高三1班" → 1
    若无数字则返回 0。
    """
    digits = re.findall(r"\d+", name)
    if digits:
        return int("".join(digits))
    return 0


def _save_workbook(wb: Workbook, title: str, subdir: str = "成绩分析") -> Path:
    """
    确保输出目录存在 → 处理文件名冲突 → 保存并返回路径。
    """
    from core.config import OUTPUT_DIR

    output_dir = OUTPUT_DIR / subdir
    output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / f"{title}.xlsx"
    if output_path.exists():
        from datetime import datetime

        timestamp = datetime.now().strftime("%H%M%S")
        output_path = output_dir / f"{title}_{timestamp}.xlsx"
        logger.info(f"文件已存在，使用新文件名: {output_path}")

    wb.save(output_path)
    logger.info(f"保存成功: {output_path}")
    return output_path


# ===================================================================
# 表头写入
# ===================================================================


def write_sheet_head(ws, direction: str, Line: StreamingMap):
    """
    写入分析报表的三行表头。

    :param ws: 工作表
    :param direction: 选科方向（"物理" / "历史"）
    :param Line: 分数线映射
    """
    subject_list = _filter_subjects(direction)

    first_line = ["班级", "科目"]
    second_line = ["", ""]
    third_line = ["", "人数"]

    for subject in subject_list:
        first_line.extend([subject, "", "", "", ""])
        for line in TOTAL_LINE:
            line_score = Line.get(f"{direction}{subject}_{line}")
            third_line.append(f"{line_score}")

    ws.append(first_line)

    for i in range(len(subject_list)):
        second_line.extend(TOTAL_LINE)
        ws.merge_cells(
            f"{get_column_letter(i * 5 + 3)}1:{get_column_letter(i * 5 + 7)}1"
        )

    ws.append(second_line)
    ws.append(third_line)
    ws.merge_cells("A1:A3")
    ws.merge_cells("B1:B2")


# ===================================================================
# 数据写入
# ===================================================================


def _write_data_rows(
    ws,
    statistics_list: List[ClassStatistics],
    direction: str,
    mode: str = "both",
):
    """
    写入各班各科的统计数据行。

    :param ws: 工作表
    :param statistics_list: 各班各科统计
    :param direction: 选科方向（"物理" / "历史"）
    :param mode: "both" → single/double 格式；"single" → 只写 single
    """
    subject_list = _filter_subjects(direction)

    # 按班级名中的数字从小到大排序
    statistics_list.sort(key=lambda s: _extract_class_number(s.name))

    for statistics in statistics_list:
        class_name = statistics.name
        count = statistics.count
        logger.debug(f"正在处理班级 {class_name} 的数据")
        data: dict = statistics.get_statistics_data()
        write_info = [statistics.name, count]
        for subject in subject_list:
            for line in TOTAL_LINE:
                obj = data.get(f"{subject}{line}", SubjectStatistics())
                if mode == "single":
                    write_info.append(obj.single)
                elif line.startswith("总分"):
                    write_info.append(obj.single)
                else:
                    write_info.append(f"{obj.single}/{obj.double}")
        ws.append(write_info)


# ===================================================================
# 数据写入（公开 API，向后兼容）
# ===================================================================


class write_data:
    """
    ``write_data(ws, statistics_list, direction)``

    写入各班各科的单双上线统计数据。
    向后兼容的 callable 包装，调用 ``_write_data_rows(..., mode="both")``。
    """

    def __call__(self, ws, statistics_list: List[ClassStatistics], direction: str):
        _write_data_rows(ws, statistics_list, direction, mode="both")

    @classmethod
    def single_only(cls, ws, statistics_list: List[ClassStatistics], direction: str):
        """写入单上线统计数据（只输出 single 值，不拼接 double）。"""
        _write_data_rows(ws, statistics_list, direction, mode="single")


# 实例化使 ``write_data(ws, ...)`` 可调用
write_data = write_data()


# ===================================================================
# 样式
# ===================================================================


def theme_excel(ws):
    """
    为工作表添加框线和居中对齐。

    :param ws: 工作表
    """
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    logger.debug(f"表格样式设置完成，共 {ws.max_row} 行，{ws.max_column} 列")


# ===================================================================
# 公开 API — 主入口
# ===================================================================


def output_statistics(
    title, statistics_list: List[ClassStatistics], direction: str, Line: StreamingMap
) -> Path:
    """
    输出单双上线统计表。

    :param title: 表名、文件名
    :param statistics_list: 各班各科单双上线统计
    :param direction: 选科方向
    :param Line: 分数线
    :return: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("无法创建工作表")

    title = title.replace("/", "") + "成绩分析"

    # Sheet 1: 单双上线
    ws.title = "单双上线"
    write_sheet_head(ws, direction, Line)
    write_data(ws, statistics_list, direction)
    theme_excel(ws)

    # Sheet 2: 单上线（格式相同，只写 single）
    ws2 = wb.create_sheet(title="单上线")
    write_sheet_head(ws2, direction, Line)
    _write_data_rows(ws2, statistics_list, direction, mode="single")
    theme_excel(ws2)

    return _save_workbook(wb, title)
