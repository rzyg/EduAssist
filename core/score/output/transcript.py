from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from core.score.models import StreamingMap, Student

# ── 模块级常量 ──────────────────────────────────────────────────────
SUBJECT_LIST_TOTAL = [
    "总分",
    "语文",
    "数学",
    "英语",
    "物理",
    "化学",
    "生物",
    "历史",
    "政治",
    "地理",
]

LINE_LIST = ["清北线", "985线", "211线", "特控线", "本科线"]


# ===================================================================
# 内部助手函数
# ===================================================================


def _get_subject_list(students):
    """
    以第一名学生的考试科目代表本班选科，返回实际存在的科目列表。
    """
    if _get_direction(students[0]):
        direction, selecting1, selecting2 = _expand_subject_abbreviation(
            students[0].selection
        )
        return ["总分", "语文", "数学", "英语", direction, selecting1, selecting2]
    else:
        return SUBJECT_LIST_TOTAL[:]


def _create_header(ws, subject_list, title):
    """写入标题行和列头（姓名 + 各科目的分数/班名/校名）。"""
    ws.append([title])

    ws["A2"] = "姓名"
    ws.merge_cells("A2:A3")
    for i in range(2, len(subject_list) * 3, 3):
        ws.cell(row=2, column=i, value=subject_list[i // 3])
        ws.cell(row=3, column=i, value="分数")
        ws.cell(row=3, column=i + 1, value="班名")
        ws.cell(row=3, column=i + 2, value="校名")
        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i + 2)


def _get_direction(student):
    """获取学生的选科方向简称（如 '物'），无选科则返回空字符串。"""
    if student.selection:
        direction, _, _ = _expand_subject_abbreviation(student.selection)
        return direction
    return ""


def _fill_cutoff_lines(ws, subject_list, students, line):
    """写入各分数线行（清北线 / 985线 / …）。"""
    direction = _get_direction(students[0]) if students[0].selection else ""
    for line_name in LINE_LIST:
        fill_list = [line_name]
        for score in subject_list:
            line_key = f"{direction}{score}_{line_name}"
            if line.has(line_key):
                fill_list.extend([line[line_key], "", ""])
            else:
                logger.warning(f"分数线数据缺失: {line_key}")
                fill_list.extend(["", "", ""])
        ws.append(fill_list)


def _fill_student_data(ws, students, line):
    """写入每个学生的成绩/排名数据以及分数线过线标记行。"""
    for index, student in enumerate(students):
        data = [student.name]
        direction = _get_direction(student)

        # 根据选科构建当前学生的科目列表
        if student.selection:
            _, selecting1, selecting2 = _expand_subject_abbreviation(student.selection)
            student_subjects = [
                "总分",
                "语文",
                "数学",
                "英语",
                direction,
                selecting1,
                selecting2,
            ]
        else:
            student_subjects = SUBJECT_LIST_TOTAL[:]

        # 小语种问题
        selecting_list = student.get_selection()
        if "小语种" in selecting_list:
            student_subjects = [
                "小语种" if item == "英语" else item for item in student_subjects
            ]

        for subject in student_subjects:
            try:
                score = student.get_data(subject)
                class_rank = student.get_data(f"{subject}班名")
                school_rank = student.get_data(f"{subject}校名")
                data.extend([score, class_rank, school_rank])
            except KeyError:
                data.extend(["", "", ""])

        ws.append(data)

        # 判断过线标记（按总分降序排列，在断层处插入过线行）
        for line_name in LINE_LIST:
            line_score = line[f"{direction}总分_{line_name}"]
            if (
                index < len(students) - 1
                and student.get_data("总分") is not None
                and students[index + 1].get_data("总分") is not None
                and student.get_data("总分")
                >= line_score
                > students[index + 1].get_data("总分")
            ):
                ws.append([f"{line_name}过线{index + 1}人"])
                row = ws.max_row
                ws.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=ws.max_column,
                )


def _apply_styles(ws):
    """为整个工作表应用对齐、边框和背景填充。"""
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="dbdbdb", end_color="dbdbdb", fill_type="solid"
    )

    # 合并表头（标题行）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws["A1"].alignment = alignment

    # 表头行（第 2-3 行）：对齐 + 边框 + 灰色填充
    for row in range(2, 4):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = thin_border
            cell.fill = header_fill

    # 数据行（第 4 行起）：对齐 + 边框，分数列添加灰色填充
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = thin_border
            # 分数列（每 3 列中的第 1 列）：2, 5, 8, 11, …
            if (col - 2) % 3 == 0 and col >= 2:
                cell.fill = header_fill


def _auto_column_width(ws):
    """根据单元格内容自动调整列宽。"""
    max_col = ws.max_column
    max_row = ws.max_row

    for j in range(2, max_col + 1):
        max_d = 1
        for i in range(4, max_row + 1):
            cell_value = ws.cell(i, j).value
            if isinstance(cell_value, str):
                d = len(cell_value.encode("utf-8"))
            else:
                d = len(str(cell_value))
            if d > max_d:
                max_d = d
        ws.column_dimensions[get_column_letter(j)].width = max_d + 1


def _setup_page_layout(ws):
    """设置页边距、缩放、居中和纸张大小。"""
    ws.page_margins = PageMargins(
        left=0.1, right=0.1, top=0.1, bottom=0.1, header=0, footer=0
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def _save_workbook(wb, title):
    """确保输出目录存在 → 处理文件名冲突 → 保存并返回路径。"""
    from core.config import OUTPUT_DIR

    output_dir = OUTPUT_DIR / "成绩单"
    output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / f"{title}.xlsx"
    if output_path.exists():
        from datetime import datetime

        timestamp = datetime.now().strftime("%H%M%S")
        output_path = output_dir / f"{title}_{timestamp}.xlsx"
        logger.info(f"文件已存在，使用新文件名: {output_path}")

    wb.save(output_path)
    logger.info(f"保存成功: {output_path}")
    return output_path


# ===================================================================
# 公开 API
# ===================================================================


def output_transcript(
    title,
    students_list: list[Student],
    Line: StreamingMap,
):
    """
    输出成绩单

    参数:
    - title: 表格标题
    - students_list: 一个列表，包含所有学生的数据
    - Line: 一个StreamingMap对象，包含分数线的数据
    """
    title = title.replace("/", "") + "成绩公示"

    # 判断一下选科，以第一名考试科目代表本班选科
    subject_list = _get_subject_list(students_list)

    wb = Workbook()
    ws = wb.active
    # 确保 ws 不为 None
    if ws is None:
        raise ValueError("无法创建工作表")

    _create_header(ws, subject_list, title)

    # 填充分数线
    _fill_cutoff_lines(ws, subject_list, students_list, Line)

    # 填充学生成绩和过线情况
    _fill_student_data(ws, students_list, Line)

    # 应用样式
    _apply_styles(ws)

    # 自动调整列宽
    _auto_column_width(ws)

    # 设置页面缩放和居中
    _setup_page_layout(ws)

    return _save_workbook(wb, title)


def _expand_subject_abbreviation(abbreviation: str) -> tuple[str, str, str]:
    """
    从选科简称还原为完整科目列表

    参数:
        abbreviation: 选科简称,如 "物化地"、"史政地"、"物化生"

    返回:
        完整科目元组,如 ("物理", "化学", "地理")

    异常:
        IndexError: 当 abbreviation 长度不足 3（选科简称应恰好3字）
    """
    # 定义简称到全称的映射
    abbreviation_map = {
        "物": "物理",
        "化": "化学",
        "生": "生物",
        "史": "历史",
        "政": "政治",
        "地": "地理",
    }

    # 逐字转换
    result: list[str] = []
    for char in abbreviation:
        if char in abbreviation_map:
            result.append(abbreviation_map[char])
        else:
            logger.warning(f"未知的选科简称字符: {char}")

    if len(result) != 3:
        logger.error(
            f"选科简称 '{abbreviation}' 转换后只有 {len(result)} 科，期望 3 科"
        )
        # 补全到 3 个，防止 IndexError
        while len(result) < 3:
            result.append("")
    return (result[0], result[1], result[2])
