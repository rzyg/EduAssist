from pathlib import Path
from loguru import logger
from core.score.models import StreamingMap
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# 加载excel文件
def loadData(path: str | Path) -> Worksheet:
    """
    加载excel文件
    :param path: 文件路径
    :return: 工作表
    """
    wb = load_workbook(filename=path)
    return wb.worksheets[0]


def getPosition(
    ws: Worksheet,
    context: str,
    start_row: int = 1,
    start_col: int = 1,
    end_row: int = 0,
    end_col: int = 0,
):
    """
    在工作表中查询某内容的坐标（包含匹配）

    :param ws: Excel工作表
    :param context: 查询内容（只要单元格包含此字符串即匹配）
    :param start_row: 起始行（默认1）
    :param start_col: 起始列（默认1）
    :param end_row: 结束行（默认 ws.max_row）
    :param end_col: 结束列（默认 ws.max_column）
    :return: [row_idx, col_idx] 或 None
    """
    # 设置默认结束范围
    if end_row == 0:
        end_row = ws.max_row
    if end_col == 0:
        end_col = ws.max_column

    # 限制搜索范围，避免扫描整个表
    for row_idx in range(start_row, min(end_row, ws.max_row) + 1):
        for col_idx in range(start_col, min(end_col, ws.max_column) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value and context in str(value):
                logger.debug(f"找到 '{context}' 在 [{row_idx}, {col_idx}]")
                return [row_idx, col_idx]

    logger.warning(f"未找到包含 '{context}' 的单元格")
    return None


def build_score_mapping(ws: Worksheet):
    """
    根据工作表构建列映射
    自动适配赋分/非赋分：优先匹配赋分列，不存在则匹配原始分列

    Args:
        ws: Excel 工作表

    Returns:
        StreamingMap 实例，包含所有字段的列号
    """
    global streaming
    mapping = StreamingMap()

    class_name = getPosition(ws, "班", 1, 1, 5)
    if class_name is None:
        raise ValueError("Excel 表中缺少班级字段")
    mapping["班级"] = class_name[1]

    name = getPosition(ws, "姓名", 1, 1, 5)
    if name is None:
        raise ValueError("Excel 表中缺少姓名字段")
    mapping["姓名"] = name[1]

    # ========== 1. 总分（自动适配赋分/原始）==========
    # getPosition 会优先查"赋分总分"，没有则查"总分"
    total_pos = getPosition(ws, "总分", 1, 2, 5)
    if total_pos is None:
        raise ValueError("Excel 表中缺少总分字段")

    total_x = total_pos[1]
    mapping["总分"] = total_x
    mapping["总分班名"] = getPosition(ws, "班", 1, total_x, 5)[1]
    mapping["总分校名"] = getPosition(ws, "校", 1, total_x, 5)[1]

    # ========== 2. 固定科目（语数英）==========
    for subject in ["语文", "数学", "英语"]:
        pos = getPosition(ws, subject, 1, 1, 5)
        if pos is None:
            raise ValueError(f"Excel 表中缺少必要字段: {subject}")

        subject_x = pos[1]
        mapping[f"{subject}"] = subject_x
        mapping[f"{subject}班名"] = getPosition(ws, "班", 1, subject_x, 5)[1]
        mapping[f"{subject}校名"] = getPosition(ws, "校", 1, subject_x, 5)[1]

    # ========== 3. 选考科目（动态匹配，存在则添加）==========
    optional_subjects = ["物理", "历史", "生物", "化学", "地理", "政治", "小语种"]
    found_subjects = []

    for subject in optional_subjects:
        pos = getPosition(ws, subject, 1, total_x, 5)
        if pos:
            subject_x = pos[1]
            mapping[f"{subject}"] = subject_x
            mapping[f"{subject}班名"] = getPosition(ws, "班", 1, subject_x, 5)[1]
            mapping[f"{subject}校名"] = getPosition(ws, "校", 1, subject_x, 5)[1]
            found_subjects.append(subject)
            logger.debug(f"找到选考科目: {subject}，列号: {subject_x}")
        else:
            logger.debug(f"未找到选考科目: {subject}，跳过")

    logger.info(f"列映射构建完成，共 {len(mapping)} 个字段，选考科目: {found_subjects}")

    # 判断分科
    if "物理" and "历史" in found_subjects:
        logger.debug("未分科")
        streaming = "false"
    elif "物理" in found_subjects and "历史" not in found_subjects:
        logger.debug("物理分科")
        streaming = "physics"
    elif "物理" not in found_subjects and "历史" in found_subjects:
        logger.warning("历史分科")
        streaming = "history"
    return mapping, streaming


def get_lines(ws: Worksheet) -> StreamingMap:
    """
    根据工作表构建分数线映射，提取所有分数线的实际数值
    Args:
        ws: Excel 工作表

    Returns:
        StreamingMap 实例，包含所有分数线的数值
    """
    mapping = StreamingMap()
    line_map = StreamingMap()
    physics_line = getPosition(ws, "物理", 1, 1)
    history_line = getPosition(ws, "历史", 1, 1)
    subject_list = [
        "总分",
        "语文",
        "数学",
        "英语",
        "物理",
        "历史",
        "生物",
        "化学",
        "地理",
        "政治",
        "小语种",
    ]
    # 判断是否分科
    if physics_line[1] == history_line[1]:
        # 未分科
        logger.debug(f"找到物理/历史列号: {physics_line}/{history_line}，判断为未分科")

        for subject in subject_list:
            mapping[f"{subject}"] = getPosition(ws, subject, 2, 1)[0]
        for line in ["清北线", "985线", "211线", "特控线", "本科线"]:
            col = getPosition(ws, line, 1, 1)[1]
            mapping[f"{line}"] = col
            # 提取分数线实际数值：遍历所有学科行，提取该列的值
            for subject in subject_list:
                row = mapping.get(f"{subject}")
                if row is not None and isinstance(row, int) and row > 0:
                    value = ws.cell(row=row, column=col).value
                    line_map[f"{subject}_{line}"] = to_float(value)

    elif physics_line[1] != history_line[1]:
        # 分科
        logger.debug(f"物理/历史列号不同: {physics_line}/{history_line}，判断为分科")
        history_col = getPosition(ws, "历史", 1, 1, 1)[1]

        # 物理方向
        subject_list.remove("历史")
        for subject in subject_list:
            mapping[f"物理{subject}"] = getPosition(
                ws, subject, 2, 1, 0, history_col - 1
            )[0]
        for line in ["清北线", "985线", "211线", "特控线", "本科线"]:
            col = getPosition(ws, line, 1, 1)[1]
            mapping[f"物理{line}"] = col
            # 提取物理方向分数线实际数值
            for subject in subject_list:
                row = mapping.get(f"物理{subject}")
                if row is not None and isinstance(row, int) and row > 0:
                    value = ws.cell(row=row, column=col).value
                    line_map[f"物理{subject}_{line}"] = to_float(value)

        # 历史方向

        subject_list.remove("物理")
        subject_list.append("历史")
        for subject in subject_list:
            mapping[f"历史{subject}"] = getPosition(ws, subject, 2, history_col)[0]
        for line in ["清北线", "985线", "211线", "特控线", "本科线"]:
            col = getPosition(ws, line, 1, history_col)[1]
            mapping[f"历史{line}"] = col
            # 提取历史方向分数线实际数值
            for subject in subject_list:
                row = mapping.get(f"历史{subject}")
                if row is not None and isinstance(row, int) and row > 0:
                    value = ws.cell(row=row, column=col).value
                    line_map[f"历史{subject}_{line}"] = to_float(value)

    elif physics_line is None or history_line is None:
        raise ValueError("Excel 表中缺少必要字段: 物理/历史")
    return line_map


def to_float(value, default=0):
    """安全转换为浮点数"""
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning(f"无法转换为浮点数: {value}")
        return default
