from typing import List, Dict
from core.score.models import StreamingMap, Student, SubjectScore
from openpyxl.worksheet.worksheet import Worksheet
from loguru import logger


def extract_score(
    score_sheet: Worksheet,
    map_list: StreamingMap,
) -> List[Student]:
    """
    处理学生成绩
    :param score_sheet: 成绩单工作表
    :param map_list: 成绩单列映射表
    :return: 学生对象列表
    """
    logger.info("开始提取学生成绩")
    student_list = []
    required_subjects = [
        "总分",
        "语文",
        "数学",
        "英语",
        "物理",
        "化学",
        "生物",
        "历史",
        "政治",
        "地理",
    ]
    # 遍历每一行数据
    for row in range(2, score_sheet.max_row + 1):
        # 获取基础信息
        student_class = score_sheet.cell(row=row, column=int(map_list["班级"])).value
        student_name = score_sheet.cell(row=row, column=int(map_list["姓名"])).value
        # 跳过空行（没有姓名的行）
        if not student_name or student_name == "姓名":
            logger.debug(f"跳过空行: {row}")
            continue

        # 构建 subjects 字典
        subjects: Dict[str, SubjectScore] = {}

        # 处理
        for subject in required_subjects:
            try:
                column = map_list[subject]
            except KeyError:
                logger.warning(f"{subject} 列未找到")
                continue
            score = float(
                score_sheet.cell(row=row, column=int(map_list[subject])).value or 0
            )
            class_rank = score_sheet.cell(
                row=row, column=int(map_list[f"{subject}班名"])
            ).value
            school_rank = score_sheet.cell(
                row=row, column=int(map_list[f"{subject}校名"])
            ).value

            # 只有成绩不为空时才添加
            if isinstance(score, float) and class_rank and school_rank:
                subjects[subject] = SubjectScore(
                    score=score, class_rank=class_rank, school_rank=school_rank
                )
            else:
                logger.warning(f"{student_class}班{student_name} 没有 {subject} 科目")
        if not subjects.get("英语"):
            logger.warning(f"{student_class}班{student_name} 没有英语科目")
            # 没有英语科目时，使用小语种代替
            small_score = float(
                score_sheet.cell(row=row, column=int(map_list["小语种"])).value or 0
            )
            small_class_rank = score_sheet.cell(
                row=row, column=int(map_list["小语种班名"])
            ).value
            small_school_rank = score_sheet.cell(
                row=row, column=int(map_list["小语种校名"])
            ).value

            if small_score and small_class_rank and small_school_rank:
                subjects["英语"] = SubjectScore(
                    score=float(small_score),
                    class_rank=small_class_rank,
                    school_rank=small_school_rank,
                )
        # 读取选科
        selection = (
            str(score_sheet.cell(row=row, column=int(map_list["选科"])).value) or ""
        )
        # 创建学生对象并添加到列表
        student = Student(
            student_class=student_class,
            name=student_name,
            subjects=subjects,
            selection=selection,
        )
        student_list.append(student)

    return student_list
