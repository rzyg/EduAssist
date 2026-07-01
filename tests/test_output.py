"""
core/score/transcript/output.py 单元测试

测试目标:
  - expand_subject_abbreviation: 选科简称展开
  - create_table: 成绩单 Excel 输出（验证文件生成与基本结构）
"""
import pytest
from openpyxl import load_workbook
from core.score.transcript.output import expand_subject_abbreviation, create_table
from core.score.models import StreamingMap


# =============================================================================
# expand_subject_abbreviation
# =============================================================================

class TestExpandSubjectAbbreviation:
    """选科简称展开"""

    def test_physics_chemistry_biology(self):
        """物化生 → 物理, 化学, 生物"""
        result = expand_subject_abbreviation("物化生")
        assert result == ("物理", "化学", "生物")

    def test_history_politics_geography(self):
        """史政地 → 历史, 政治, 地理"""
        result = expand_subject_abbreviation("史政地")
        assert result == ("历史", "政治", "地理")

    def test_physics_chemistry_geography(self):
        """物化地 → 物理, 化学, 地理"""
        result = expand_subject_abbreviation("物化地")
        assert result == ("物理", "化学", "地理")

    def test_return_type_is_tuple(self):
        """返回值是元组"""
        result = expand_subject_abbreviation("物化生")
        assert isinstance(result, tuple)
        assert len(result) == 3

    def test_unknown_char_warns(self):
        """未知字符不会导致崩溃，但结果可能为空"""
        result = expand_subject_abbreviation("XYZ")
        # 未知字符被跳过
        assert len(result) == 3  # 会被 padding 为 3 个
        # 所有有效字符都被映射，无效的被忽略

    def test_empty_string(self):
        """空字符串不会崩溃"""
        result = expand_subject_abbreviation("")
        assert len(result) == 3
        assert result == ("", "", "")


# =============================================================================
# create_table
# =============================================================================

class TestCreateTable:
    """成绩单 Excel 输出"""

    def test_output_file_created(self, tmp_path):
        """
        基本场景：3 名无选科学生，输出文件被创建且包含基本内容。
        """
        from core.score.models import Student, SubjectScore
        students = [
            Student("一班", "张三", {
                "总分": SubjectScore(600, 1, 2),
                "语文": SubjectScore(130, 1, 3),
                "数学": SubjectScore(142, 1, 2),
                "英语": SubjectScore(128, 1, 2),
            }, ""),
            Student("一班", "李四", {
                "总分": SubjectScore(550, 2, 5),
                "语文": SubjectScore(120, 2, 5),
                "数学": SubjectScore(135, 2, 5),
                "英语": SubjectScore(118, 2, 6),
            }, ""),
            Student("一班", "王五", {
                "总分": SubjectScore(500, 3, 8),
                "语文": SubjectScore(110, 3, 8),
                "数学": SubjectScore(125, 3, 8),
                "英语": SubjectScore(105, 3, 10),
            }, ""),
        ]

        from core.score.models import StreamingMap
        line_map = StreamingMap()
        line_map.update({
            "总分_清北线": 600, "总分_985线": 550, "总分_211线": 500,
            "总分_特控线": 450, "总分_本科线": 400,
            "语文_清北线": 130, "语文_985线": 120, "语文_211线": 110,
            "语文_特控线": 100, "语文_本科线": 90,
            "数学_清北线": 140, "数学_985线": 130, "数学_211线": 120,
            "数学_特控线": 110, "数学_本科线": 100,
            "英语_清北线": 130, "英语_985线": 120, "英语_211线": 110,
            "英语_特控线": 100, "英语_本科线": 90,
        })

        output_path = create_table("测试考试", students, line_map)
        assert output_path.exists()
        wb = load_workbook(output_path)
        ws = wb.active

        # create_table 输出中含标题、表头、分数线和学生数据
        assert "测试考试成绩公示" in str(ws["A1"].value)
        assert ws["A2"].value == "姓名"
        # 遍历 A 列找到所有学生姓名
        all_cells_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "张三" in all_cells_a
        assert "李四" in all_cells_a
        assert "王五" in all_cells_a

    def test_with_no_selection_students(self, tmp_path):
        """学生没有选科信息时也能生成"""
        from core.score.models import Student, SubjectScore, StreamingMap
        students = [
            Student("一班", "赵六", {
                "总分": SubjectScore(600, 1, 1),
                "语文": SubjectScore(130, 1, 2),
            }, "")
        ]
        line_map = StreamingMap()
        line_map.update({
            "总分_清北线": 600, "总分_985线": 550, "总分_211线": 500,
            "总分_特控线": 450, "总分_本科线": 400,
            "语文_清北线": 130, "语文_985线": 120, "语文_211线": 110,
            "语文_特控线": 100, "语文_本科线": 90,
        })
        output_path = create_table("无选科测试", students, line_map)
        assert output_path.exists()
        wb = load_workbook(output_path)
        ws = wb.active
        assert "无选科测试成绩公示" in str(ws["A1"].value)
        names = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "赵六" in names

    def test_with_single_student(self, tmp_path):
        """单名学生也能生成"""
        from core.score.models import Student, SubjectScore, StreamingMap
        single = [
            Student("一班", "独苗", {
                "总分": SubjectScore(600, 1, 1),
                "语文": SubjectScore(130, 1, 2),
            }, "")
        ]
        line_map = StreamingMap()
        line_map.update({
            "总分_清北线": 600, "总分_985线": 550, "总分_211线": 500,
            "总分_特控线": 450, "总分_本科线": 400,
            "语文_清北线": 130, "语文_985线": 120, "语文_211线": 110,
            "语文_特控线": 100, "语文_本科线": 90,
        })
        output_path = create_table("单人测试", single, line_map)
        assert output_path.exists()
        wb = load_workbook(output_path)
        ws = wb.active
        names = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "独苗" in names
