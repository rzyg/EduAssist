"""
core/allowance/check 坐班签到统计单元测试

测试目标:
  - 签到表解析（6 行结构、合并单元格、时间/'-'/空处理、日期列头）
  - 各时段有效性判定规则（正常上午/下午、离校下午、晚上）
  - 结合假期表的统计（正常日/离校日/返校日/假期）
  - 统计 Excel 输出（标题/表头/数据/时间戳）
  - POST /attendance_statistics 路由
"""
import io
import os
from datetime import date, datetime, time
from pathlib import Path
from unittest import mock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from core.allowance.check import output as O
from core.allowance.check.load_data import load_teachers, parse_date_header, parse_time
from core.allowance.check.modules import Check, Teacher, Workday
from core.allowance.check.statistics import (
    AFTERNOON,
    EVENING,
    LEAVE_AFTERNOON,
    MORNING,
    build_holiday_map,
    compute_attendance,
    is_valid_check,
)
from core.route.allowance import router

# =============================================================================
# 辅助：构造签到表
# =============================================================================

HEADERS = [
    "姓名", "部门",
    "周一 26-03-09", "周二 26-03-10", "周三 26-03-11",
    "周四 26-03-12", "周五 26-03-13", "周六 26-03-14", "周日 26-03-15",
]

# 与用户提供的示例一致（NaN 用 None 表示空单元格）
SAMPLE_ROWS = [
    ["曹雪伟", "高一1部", "07:43", "07:41", "07:26", "07:36", "07:23", None, "18:44"],
    [None, None, "11:32", "11:36", "11:46", "11:04", "11:21", None, "21:04"],
    [None, None, "14:03", "13:45", "14:01", "13:49", "14:37", None, None],
    [None, None, "17:39", "17:20", "17:22", "17:54", "15:43", None, None],
    [None, None, "18:38", "-", "18:49", "18:47", None, None, None],
    [None, None, "21:45", "-", "21:09", "21:06", None, None, None],
]


def build_sheet(rows=SAMPLE_ROWS):
    """构造签到表 Worksheet（内存）"""
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col, value=value)
    return ws


def build_xlsx_bytes(rows=SAMPLE_ROWS) -> bytes:
    """构造签到表 xlsx 文件字节"""
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col, value=value)
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# 解析
# =============================================================================

class TestParse:
    def test_parse_time(self):
        assert parse_time("07:43") == time(7, 43)
        assert parse_time("-") is None
        assert parse_time(None) is None
        assert parse_time("") is None
        assert parse_time("NaN") is None

    def test_parse_date_header(self):
        assert parse_date_header("周一 26-03-09") == date(2026, 3, 9)
        assert parse_date_header("周日 26-03-15") == date(2026, 3, 15)
        assert parse_date_header("2026-03-09") == date(2026, 3, 9)
        assert parse_date_header("姓名") is None

    def test_parse_date_header_newline(self):
        """星期与日期之间为回车/换行（实际表中常为单元格内换行）"""
        assert parse_date_header("周一\n26-03-09") == date(2026, 3, 9)
        assert parse_date_header("周一\r\n26-03-09") == date(2026, 3, 9)
        assert parse_date_header("26-03-09\n周一") == date(2026, 3, 9)

    def test_parse_date_header_datetime(self):
        """openpyxl 把日期列头读成 datetime 时直接取日期"""
        assert parse_date_header(datetime(2026, 3, 9)) == date(2026, 3, 9)
        assert parse_date_header("2026/3/9") == date(2026, 3, 9)
        assert parse_date_header("2026.03.09") == date(2026, 3, 9)

    def test_load_teachers_structure(self):
        ws = build_sheet()
        teachers = load_teachers(ws)
        assert len(teachers) == 1
        teacher = teachers[0]
        assert teacher.name == "曹雪伟"
        assert teacher.department == "高一1部"
        assert len(teacher.workdays) == 7
        # 周一：三时段均有打卡
        monday = teacher.workdays[0]
        assert monday.date.date() == date(2026, 3, 9)
        assert monday.morning.time_in == time(7, 43)
        assert monday.morning.time_out == time(11, 32)
        assert monday.afternoon.time_in == time(14, 3)
        assert monday.afternoon.time_out == time(17, 39)
        assert monday.evening.time_in == time(18, 38)
        assert monday.evening.time_out == time(21, 45)
        # 周二晚上为 '-' → 无打卡
        assert teacher.workdays[1].evening.time_in is None
        # 周日：返校日数据在前两行（morning 位）
        sunday = teacher.workdays[6]
        assert sunday.morning.time_in == time(18, 44)
        assert sunday.morning.time_out == time(21, 4)

    def test_load_teachers_no_date_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="姓名")
        with pytest.raises(ValueError, match="日期列"):
            load_teachers(ws)


# =============================================================================
# 时段规则判定
# =============================================================================

def make_check(in_time: str, out_time: str) -> Check:
    return Check(
        time_in=datetime(2026, 3, 9, int(in_time[:2]), int(in_time[3:])),
        time_out=datetime(2026, 3, 9, int(out_time[:2]), int(out_time[3:])),
    )


class TestSlotRules:
    def test_morning(self):
        assert is_valid_check(make_check("07:43", "11:32"), MORNING)
        assert not is_valid_check(make_check("08:51", "11:32"), MORNING)
        assert not is_valid_check(make_check("07:43", "10:59"), MORNING)

    def test_afternoon(self):
        assert is_valid_check(make_check("14:03", "17:39"), AFTERNOON)
        assert not is_valid_check(make_check("14:03", "16:59"), AFTERNOON)
        assert not is_valid_check(make_check("15:01", "17:39"), AFTERNOON)

    def test_leave_afternoon(self):
        assert is_valid_check(make_check("14:03", "15:43"), LEAVE_AFTERNOON)
        assert not is_valid_check(make_check("14:03", "15:00"), LEAVE_AFTERNOON)

    def test_evening(self):
        assert is_valid_check(make_check("18:38", "21:45"), EVENING)
        assert not is_valid_check(make_check("18:34", "21:45"), EVENING)
        assert not is_valid_check(make_check("18:38", "20:59"), EVENING)

    def test_missing_check(self):
        assert not is_valid_check(None, MORNING)
        assert not is_valid_check(Check(), MORNING)


# =============================================================================
# 统计（结合假期表）
# =============================================================================

class TestComputeAttendance:
    def test_sample_week_with_return_day(self):
        """示例数据：7 天共 13 次，周日为返校日按晚上规则计 1 次"""
        teachers = load_teachers(build_sheet())
        holiday_map = build_holiday_map([
            {"year": 2026, "month_day": "03-15", "holiday": True,
             "leave_school": False, "return_school": True},
        ])
        compute_attendance(teachers, holiday_map)
        teacher = teachers[0]
        assert teacher.count == 13  # 3+2+3+3+1+0+1
        assert teacher.workdays[6].return_school is True

    def test_leave_school_day(self):
        """离校日：上午 + 离校下午，晚上不统计"""
        rows = [
            ["李四", "高一2部", "07:30", None, None, None, None, None, None],
            [None, None, "11:10", None, None, None, None, None, None],
            [None, None, "14:00", None, None, None, None, None, None],
            [None, None, "15:20", None, None, None, None, None, None],
            [None, None, "18:40", None, None, None, None, None, None],
            [None, None, "21:10", None, None, None, None, None, None],
        ]
        teachers = load_teachers(build_sheet(rows))
        holiday_map = build_holiday_map([
            {"year": 2026, "month_day": "03-09", "holiday": False,
             "leave_school": True, "return_school": False},
        ])
        compute_attendance(teachers, holiday_map)
        assert teachers[0].count == 2  # 上午 + 离校下午

    def test_holiday_not_counted(self):
        """假期（非返校日）不统计"""
        rows = [
            ["王五", "高一3部", "07:30", None, None, None, None, None, None],
            [None, None, "11:10", None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None],
        ]
        teachers = load_teachers(build_sheet(rows))
        holiday_map = build_holiday_map([
            {"year": 2026, "month_day": "03-09", "holiday": True,
             "leave_school": False, "return_school": False},
        ])
        compute_attendance(teachers, holiday_map)
        assert teachers[0].count == 0

    def test_normal_weekend_counted_when_has_records(self):
        """普通周末无假期标记时按正常规则统计（示例周六无打卡为 0）"""
        teachers = load_teachers(build_sheet())
        compute_attendance(teachers, {})
        assert teachers[0].workdays[5].date.date() == date(2026, 3, 14)
        assert teachers[0].count == 12  # 13 - 周日返校日 1 次（无标记时按正常规则周日不计）


# =============================================================================
# 输出 Excel
# =============================================================================

class TestOutput:
    def test_output_statistics(self, tmp_path):
        """标题/表头/数据/按部门分组正确"""
        teachers = [
            Teacher(name="曹雪伟", department="高一1部", workdays=[], count=13),
            Teacher(name="李四", department="高一2部", workdays=[], count=2),
        ]
        orig = O.OUTPUT_SUBDIR
        O.OUTPUT_SUBDIR = tmp_path
        try:
            paths = O.output_statistics(teachers, (date(2026, 3, 9), date(2026, 3, 15)))
        finally:
            O.OUTPUT_SUBDIR = orig
        assert len(paths) == 2
        g1 = next(p for p in paths if "高一1部" in p.name)
        ws = load_workbook(g1).active
        assert "签到统计" in ws.cell(row=1, column=1).value
        assert ws.cell(row=2, column=1).value == "姓名"
        assert ws.cell(row=2, column=2).value == "签到次数"
        assert ws.cell(row=3, column=1).value == "曹雪伟"
        assert ws.cell(row=3, column=2).value == 13

    def test_unique_path_timestamp(self, tmp_path):
        """同名文件追加时间戳"""
        first = O._unique_path(tmp_path, "统计")
        wb = Workbook()
        wb.active.cell(row=1, column=1, value="x")
        wb.save(first)
        second = O._unique_path(tmp_path, "统计")
        assert second != first
        assert "_" in second.name


# =============================================================================
# 路由
# =============================================================================

class TestAttendanceRoute:
    def test_post_statistics(self, tmp_path):
        """上传签到表 → 200 返回输出文件路径（含返校日处理）"""
        app = FastAPI()
        app.include_router(router)
        holiday_records = [
            {"year": 2026, "month_day": "03-15", "holiday": True,
             "leave_school": False, "return_school": True},
        ]
        with mock.patch("core.route.allowance.get_holiday_data", return_value=holiday_records):
            with mock.patch("core.allowance.check.output.OUTPUT_SUBDIR", new=Path(tmp_path)):
                client = TestClient(app)
                resp = client.post(
                    "/api/v1/allowance/attendance_statistics",
                    files={"sheet": ("signin.xlsx", build_xlsx_bytes(),
                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
        assert resp.status_code == 200, resp.text
        out = resp.json()["output_path"]
        assert len(out) == 1 and os.path.exists(out[0])
        ws = load_workbook(out[0]).active
        assert ws.cell(row=3, column=1).value == "曹雪伟"
        assert ws.cell(row=3, column=2).value == 13
