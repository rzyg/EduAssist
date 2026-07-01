"""
core/score/map.py 单元测试

测试目标:
  - to_float: 安全类型转换
  - loadData: 文件加载
  - getPosition: 表头搜索
  - build_score_mapping: 列映射构建
  - get_lines: 分数线提取（分科/不分科）
"""
import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from core.score.map import to_float, loadData, getPosition, build_score_mapping, get_lines
from core.score.models import StreamingMap


# =============================================================================
# to_float
# =============================================================================

class TestToFloat:
    """安全类型转换"""

    def test_int(self):
        assert to_float(42) == 42.0

    def test_float(self):
        assert to_float(3.14) == 3.14

    def test_numeric_string(self):
        assert to_float("95.5") == 95.5

    def test_none(self):
        """None 返回默认值 0"""
        assert to_float(None) == 0.0

    def test_non_numeric_string(self):
        """无法转换的字符串返回默认值 0"""
        assert to_float("abc") == 0.0

    def test_custom_default(self):
        """可指定自定义默认值"""
        assert to_float(None, default=-1) == -1.0


# =============================================================================
# loadData
# =============================================================================

class TestLoadData:
    """Excel 文件加载"""

    def test_load_existing_file(self, tmp_path: Path):
        """能成功加载一个有效的 .xlsx 文件"""
        path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(path))
        ws = loadData(str(path))
        assert ws is not None

    def test_load_missing_file(self, tmp_path: Path):
        """不存在的文件抛 FileNotFoundError"""
        path = tmp_path / "不存在.xlsx"
        with pytest.raises(FileNotFoundError):
            loadData(str(path))


# =============================================================================
# getPosition
# =============================================================================

class TestGetPosition:
    """表头/单元格内容搜索"""

    def test_find_exact(self, score_worksheet: Worksheet):
        """精确匹配找到单元格"""
        pos = getPosition(score_worksheet, "姓名", 1, 1, 5, 10)
        assert pos == [1, 2]  # 第 1 行第 2 列

    def test_find_partial(self, score_worksheet: Worksheet):
        """包含匹配也能找到"""
        pos = getPosition(score_worksheet, "班", 1, 1, 5)
        # "班级"在第1列，或"语文班名"在第5列——取先找到的
        assert pos is not None
        assert pos[0] == 1  # 第 1 行

    def test_not_found(self, score_worksheet: Worksheet):
        """不存在的字符串返回 None"""
        pos = getPosition(score_worksheet, "外星语", 1, 1, 5)
        assert pos is None

    def test_search_range(self, score_worksheet: Worksheet):
        """指定搜索行列范围有效"""
        # 只在第 1 列第 1 行范围内搜索"语文"，不会找到
        pos = getPosition(score_worksheet, "语文", 1, 1, 1, 1)
        assert pos is None

    def test_data_row(self, score_worksheet: Worksheet):
        """能搜索到数据行内容（不只是表头）"""
        pos = getPosition(score_worksheet, "张三", 2, 1)
        assert pos == [2, 2]  # 第 2 行第 2 列


# =============================================================================
# build_score_mapping
# =============================================================================

class TestBuildScoreMapping:
    """自动列映射构建"""

    def test_basic_mapping(self, score_worksheet: Worksheet):
        """标准成绩单正确构建映射"""
        mapping = build_score_mapping(score_worksheet)
        assert mapping["班级"] == 1
        assert mapping["姓名"] == 2
        assert mapping["选科"] == 3
        assert mapping["语文"] == 4
        assert mapping["数学"] == 7
        assert mapping["英语"] == 10
        assert mapping["总分"] == 13
        # 验证排名列也存在
        assert mapping["语文班名"] == 5
        assert mapping["语文校名"] == 6
        assert mapping["总分班名"] == 14
        assert mapping["总分校名"] == 15
        # 选考科目在总分之后
        assert mapping["物理"] == 16
        assert mapping["化学"] == 19
        assert mapping["生物"] == 22

    def test_missing_class(self):
        """缺少「班级」列时抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "姓名"
        ws["B1"] = "语文"
        with pytest.raises(ValueError, match="缺少班级"):
            build_score_mapping(ws)

    def test_missing_name(self):
        """缺少「姓名」列时抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "班级"
        with pytest.raises(ValueError, match="缺少姓名"):
            build_score_mapping(ws)

    def test_missing_subject(self):
        """缺少必要科目（语文/数学/英语）时抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "班级"
        ws["B1"] = "姓名"
        with pytest.raises(ValueError, match="缺少"):
            build_score_mapping(ws)

    def test_optional_subjects(self, score_worksheet: Worksheet):
        """选考科目（物理/化学/生物）在总分列之后被自动检测到"""
        mapping = build_score_mapping(score_worksheet)
        assert mapping["物理"] == 16
        assert mapping["化学"] == 19
        assert mapping["生物"] == 22

    def test_no_selection_field(self, score_worksheet_no_selection: Worksheet):
        """「选科」字段缺失时不会崩溃，只是警告"""
        mapping = build_score_mapping(score_worksheet_no_selection)
        assert mapping["班级"] == 1
        assert mapping["姓名"] == 2
        # 选科不应该在 mapping 中
        assert "选科" not in mapping


# =============================================================================
# get_lines
# =============================================================================

class TestGetLines:
    """分数线提取"""

    def test_unsplit(self, line_worksheet_unsplit: Worksheet):
        """不分科格式正确提取分数线数值"""
        lines = get_lines(line_worksheet_unsplit)
        assert lines["总分_清北线"] == 600
        assert lines["总分_985线"] == 550
        assert lines["语文_清北线"] == 130
        assert lines["数学_211线"] == 120
        assert lines["英语_本科线"] == 90

    def test_split(self, line_worksheet_split: Worksheet):
        """分科格式正确提取物理/历史方向分数线"""
        lines = get_lines(line_worksheet_split)
        assert lines["物理总分_清北线"] == 600
        assert lines["物理总分_985线"] == 550
        assert lines["物理语文_清北线"] == 130
        assert lines["历史总分_清北线"] == 580
        assert lines["历史总分_特控线"] == 430

    def test_missing_physics(self):
        """缺少「物理」行时抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "科目"
        ws["A2"] = "语文"
        ws["B1"] = "清北线"
        ws["B2"] = 600
        with pytest.raises(ValueError, match="物理/历史"):
            get_lines(ws)

    def test_split_detection(self, line_worksheet_split: Worksheet):
        """分科判断逻辑：物理/历史列号不同时走分科分支"""
        lines = get_lines(line_worksheet_split)
        # 物理侧有数据
        assert "物理总分_清北线" in lines
        # 历史侧有数据
        assert "历史总分_清北线" in lines
