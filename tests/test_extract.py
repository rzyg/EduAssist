"""
core/score/transcript/extract.py 单元测试

测试目标:
  - extract_score: 从 Excel 工作表 + StreamingMap 提取学生列表
  - 小语种后备逻辑：当英语缺失时使用小语种替代
  - 空行 / 空数据 的健壮性
"""

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from core.score.extract import extract_score
from core.score.models import StreamingMap


# =============================================================================
# 辅助函数：快速构建一个简单成绩单
# =============================================================================


def make_worksheet(headers: list[str], data: list[list]) -> Worksheet:
    """用表头和行数据创建一个内存工作表"""
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(data, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    return ws


# =============================================================================
# extract_score
# =============================================================================


class TestExtractScore:
    """成绩提取核心逻辑"""

    @pytest.fixture
    def basic_ws(self) -> Worksheet:
        """3 名学生，语数英 + 物化生"""
        headers = [
            "班级",
            "姓名",
            "选科",
            "语文",
            "语文班名",
            "语文校名",
            "数学",
            "数学班名",
            "数学校名",
            "英语",
            "英语班名",
            "英语校名",
            "总分",
            "总分班名",
            "总分校名",
        ]
        data = [
            ["一班", "甲", "物化生", 130, 1, 3, 142, 1, 2, 128, 1, 2, 600, 1, 2],
            ["一班", "乙", "物化生", 120, 2, 5, 135, 2, 5, 118, 2, 6, 550, 2, 5],
        ]
        return make_worksheet(headers, data)

    @pytest.fixture
    def basic_map(self) -> StreamingMap:
        m = StreamingMap()
        m.update(
            {
                "班级": 1,
                "姓名": 2,
                "选科": 3,
                "语文": 4,
                "语文班名": 5,
                "语文校名": 6,
                "数学": 7,
                "数学班名": 8,
                "数学校名": 9,
                "英语": 10,
                "英语班名": 11,
                "英语校名": 12,
                "总分": 13,
                "总分班名": 14,
                "总分校名": 15,
            }
        )
        return m

    def test_basic_extraction(self, basic_ws, basic_map):
        """能正确提取学生姓名和成绩"""
        students = extract_score(basic_ws, basic_map)
        assert len(students) == 2
        assert students[0].name == "甲"
        assert students[0].get_data("语文") == 130
        assert students[0].get_data("数学班名") == 1
        assert students[1].name == "乙"
        assert students[1].get_data("总分校名") == 5

    def test_skip_header_row(self, basic_ws, basic_map):
        """数据行中如果有 '姓名' 作为值（重复表头），应该跳过"""
        # 手动添加一行 '姓名' 到数据中
        ws = basic_ws
        row = ws.max_row + 1
        ws.cell(row=row, column=2, value="姓名")
        students = extract_score(ws, basic_map)
        # '姓名' 行应被跳过
        names = [s.name for s in students]
        assert "姓名" not in names

    def test_skip_empty_row(self, basic_ws, basic_map):
        """姓名为空的行应被跳过"""
        ws = basic_ws
        row = ws.max_row + 1
        ws.cell(row=row, column=2, value=None)  # 空姓名
        students = extract_score(ws, basic_map)
        assert len(students) == 2  # 没有增加

    def test_selection_preserved(self, basic_ws, basic_map):
        """选科字段被正确传递"""
        students = extract_score(basic_ws, basic_map)
        assert students[0].selection == "物化生"

    def test_missing_subject_does_not_crash(self):
        """映射中有某科目但 Excel 数据行为空，不崩溃"""
        ws = make_worksheet(
            [
                "班级",
                "姓名",
                "选科",
                "语文",
                "语文班名",
                "语文校名",
                "小语种",
                "小语种班名",
                "小语种校名",
                "总分",
                "总分班名",
                "总分校名",
            ],
            [["一班", "甲", "日", None, None, None, 95, 1, 1, 500, 1, 1]],
        )
        m = StreamingMap()
        m.update(
            {
                "班级": 1,
                "姓名": 2,
                "选科": 3,
                "语文": 4,
                "语文班名": 5,
                "语文校名": 6,
                "小语种": 7,
                "小语种班名": 8,
                "小语种校名": 9,
                "总分": 10,
                "总分班名": 11,
                "总分校名": 12,
            }
        )
        students = extract_score(ws, m)
        # 没有语文成绩，但不应崩溃
        assert students[0].name == "甲"
        # 英语后备：小语种 95 填充为英语；总分 500
        assert len(students[0].subjects) == 2


class TestSmallLanguageFallback:
    """英语科目缺失时用小语种替代"""

    @pytest.fixture
    def ws_no_english(self) -> Worksheet:
        """
        没有英语列，但有 小语种 / 小语种班名 / 小语种校名
        """
        headers = [
            "班级",
            "姓名",
            "选科",
            "语文",
            "语文班名",
            "语文校名",
            "数学",
            "数学班名",
            "数学校名",
            "小语种",
            "小语种班名",
            "小语种校名",
            "总分",
            "总分班名",
            "总分校名",
        ]
        data = [
            ["一班", "丙", "日", 110, 3, 8, 125, 3, 8, 95, 1, 3, 500, 1, 1],
        ]
        return make_worksheet(headers, data)

    @pytest.fixture
    def map_no_english(self) -> StreamingMap:
        m = StreamingMap()
        m.update(
            {
                "班级": 1,
                "姓名": 2,
                "选科": 3,
                "语文": 4,
                "语文班名": 5,
                "语文校名": 6,
                "数学": 7,
                "数学班名": 8,
                "数学校名": 9,
                "小语种": 10,
                "小语种班名": 11,
                "小语种校名": 12,
                "总分": 13,
                "总分班名": 14,
                "总分校名": 15,
            }
        )
        return m

    def test_english_exists_not_overridden(self):
        """
        英语列存在（且成绩为 0）时，不触发小语种后备。
        """
        headers = [
            "班级",
            "姓名",
            "选科",
            "英语",
            "英语班名",
            "英语校名",
            "小语种",
            "小语种班名",
            "小语种校名",
            "总分",
            "总分班名",
            "总分校名",
        ]
        data = [
            ["一班", "丁", "", 0, 1, 1, 95, 1, 3, 500, 1, 1],
        ]
        ws = make_worksheet(headers, data)
        m = StreamingMap()
        m.update(
            {
                "班级": 1,
                "姓名": 2,
                "选科": 3,
                "英语": 4,
                "英语班名": 5,
                "英语校名": 6,
                "小语种": 7,
                "小语种班名": 8,
                "小语种校名": 9,
                "总分": 10,
                "总分班名": 11,
                "总分校名": 12,
            }
        )
        students = extract_score(ws, m)
        # 英语为 0，不应被替换为小语种的 95
        assert students[0].get_data("英语") == 0.0
