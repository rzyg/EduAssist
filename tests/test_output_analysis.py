"""
core/score/output/analysis.py 单元测试

测试目标:
  - output_statistics: 分析报表 Excel 输出
  - write_sheet_head: 三行表头
  - write_data: 数据行
"""
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from core.score.output.analysis import (
    output_statistics,
    write_sheet_head,
    write_data,
    _filter_subjects,
    _extract_class_number,
)
from core.score.models import ClassStatistics, StreamingMap, SubjectStatistics


class TestWriteSheetHead:
    """分析报表表头"""

    def test_physics_direction_header(self):
        """物理方向：表头不包含历史"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        line_map = StreamingMap()
        line_map.update({
            "物理总分_清北线": 600, "物理总分_985线": 550,
            "物理总分_211线": 500, "物理总分_特控线": 450, "物理总分_本科线": 400,
            "物理语文_清北线": 130, "物理语文_985线": 120,
        })

        write_sheet_head(ws, "物理", line_map)

        # Row 1: 班级 | 科目 | 总分 (merged 5 cols) | 语文 (merged 5 cols) | ...
        # 物理方向不包含"历史"
        row1_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert ws["A1"].value == "班级"
        assert ws["B1"].value == "科目"
        # 总分应该在 C 列开始（第 1 个 subject）
        subject_headers = [v for v in row1_values if v is not None and v not in ["班级", "科目", ""]]
        assert "总分" in subject_headers
        assert "历史" not in subject_headers

    def test_history_direction_header(self):
        """历史方向：表头不包含物理"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        line_map = StreamingMap()
        line_map.update({
            "历史总分_清北线": 580, "历史总分_985线": 530,
        })

        write_sheet_head(ws, "历史", line_map)
        row1_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "物理" not in row1_values


class TestWriteData:
    """数据行写入"""

    @staticmethod
    def _init_all_subjects(stats: ClassStatistics, direction: str):
        """为所有 subject×line 组合 init_subject（模拟 run_statistics 的行为）"""
        subjects = ["总分", "语文", "数学", "英语", "物理", "化学", "生物", "政治", "地理"]
        if direction == "物理":
            subjects = [s for s in subjects if s != "历史"]
        elif direction == "历史":
            subjects = [s for s in subjects if s != "物理"]
        lines = ["清北线", "985线", "211线", "特控线", "本科线"]
        for subject in subjects:
            for line in lines:
                stats.init_subject(f"{subject}{line}")

    @pytest.fixture
    def stats(self):
        """一个班级的统计数据（所有 subject×line 均已 init）"""
        stats = ClassStatistics("高三1班", 50)
        self._init_all_subjects(stats, "物理")
        stats.increment_single("总分清北线")  # 1 人单上线
        stats.increment_single("语文清北线")  # 1 人单上线
        stats.increment_double("语文清北线")  # 1 人双上线
        stats.increment_single("数学985线")  # 1 人单上线
        return stats

    def test_write_data_format(self, stats):
        """数据行包含班级名、人数、科目统计数据"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        from core.score.models import StreamingMap
        line_map = StreamingMap()
        line_map.update({
            "物理总分_清北线": 600, "物理总分_985线": 550,
            "物理总分_211线": 500, "物理总分_特控线": 450, "物理总分_本科线": 400,
            "物理语文_清北线": 130, "物理语文_985线": 120,
            "物理语文_211线": 110, "物理语文_特控线": 100, "物理语文_本科线": 90,
            "物理数学_清北线": 140, "物理数学_985线": 130,
        })
        write_sheet_head(ws, "物理", line_map)
        write_data(ws, [stats], "物理")

        # Row 4 是第一行数据（表头占 3 行）
        data_row = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        assert data_row[0] == "高三1班"
        assert data_row[1] == 50  # 人数
        all_str = " ".join(str(v) for v in data_row)
        assert "1" in str(data_row[2])  # 总分清北线 single=1
        assert "1/1" in all_str       # 语文清北线 single=1, double=1
        assert "1/0" in all_str       # 数学985线 single=1, double=0

    def test_write_data_sorting(self):
        """班级按数字从小到大排序输出"""
        from openpyxl import Workbook

        # 创建乱序的班级列表
        stats_list = []
        for name in ["2510班", "2509班", "2511班"]:
            s = ClassStatistics(name, 40)
            self._init_all_subjects(s, "物理")
            s.increment_single("总分清北线")
            stats_list.append(s)

        wb = Workbook()
        ws = wb.active
        write_data(ws, stats_list, "物理")

        # 直接调用 write_data 时数据从第 1 行开始，应按 2509 < 2510 < 2511 排列
        class_names = [ws.cell(row=r, column=1).value for r in range(1, 4)]
        assert class_names == ["2509班", "2510班", "2511班"]

    def test_missing_subject_does_not_crash(self):
        """没有初始化小语种等科目时，不崩溃且正常输出"""
        from openpyxl import Workbook

        # 只初始化部分科目（故意不包含小语种）
        stats = ClassStatistics("测试班", 30)
        stats.init_subject("总分清北线")
        stats.increment_single("总分清北线")

        wb = Workbook()
        ws = wb.active
        # 不应抛出 KeyError
        write_data(ws, [stats], "物理")
        data_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert data_row[0] == "测试班"
        assert data_row[1] == 30


class TestOutputStatistics:
    """完整分析报表输出"""

    @staticmethod
    def _init_all_subjects(stats: ClassStatistics, direction: str):
        subjects = ["总分", "语文", "数学", "英语", "物理", "化学", "生物", "政治", "地理"]
        if direction == "物理":
            subjects = [s for s in subjects if s != "历史"]
        lines = ["清北线", "985线", "211线", "特控线", "本科线"]
        for subject in subjects:
            for line in lines:
                stats.init_subject(f"{subject}{line}")

    def test_basic_output(self, tmp_path):
        """生成包含统计数据的 Excel 文件"""
        stats = ClassStatistics("高三1班", 50)
        self._init_all_subjects(stats, "物理")
        stats.increment_single("总分清北线")
        stats.increment_single("语文清北线")
        stats.increment_double("语文清北线")

        line_map = StreamingMap()
        line_map.update({
            "物理总分_清北线": 600, "物理总分_985线": 550,
            "物理总分_211线": 500, "物理总分_特控线": 450, "物理总分_本科线": 400,
            "物理语文_清北线": 130, "物理语文_985线": 120,
            "物理语文_211线": 110, "物理语文_特控线": 100, "物理语文_本科线": 90,
        })

        path = output_statistics("测试分析", [stats], "物理", line_map)
        assert path.exists()
        # 文件名包含标题
        assert "测试分析成绩分析" in path.name

        wb = load_workbook(path)
        ws = wb.active
        # 表头第一行是 "班级" | "科目" | subject1 | ...
        assert ws["A1"].value == "班级"

        # 验证数据行存在
        names = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "高三1班" in names

    def test_output_classes_sorted(self, tmp_path):
        """多个班级按数字从小到大排序输出"""
        line_map = StreamingMap()
        line_map.update({
            "物理总分_清北线": 600, "物理总分_985线": 550,
            "物理总分_211线": 500, "物理总分_特控线": 450, "物理总分_本科线": 400,
            "物理语文_清北线": 130, "物理语文_985线": 120,
        })

        stats_list = []
        for name in ["高三10班", "高三2班", "高三1班"]:
            s = ClassStatistics(name, 40)
            self._init_all_subjects(s, "物理")
            s.increment_single("总分清北线")
            stats_list.append(s)

        path = output_statistics("排序测试", stats_list, "物理", line_map)
        wb = load_workbook(path)
        ws = wb.active
        class_names = [ws.cell(row=r, column=1).value for r in range(4, 7)]
        assert class_names == ["高三1班", "高三2班", "高三10班"]


# =============================================================================
# _extract_class_number
# =============================================================================


class TestExtractClassNumber:
    """班级名数字提取"""

    def test_pure_number(self):
        """纯数字班级名"""
        assert _extract_class_number("2510") == 2510

    def test_number_with_ban_suffix(self):
        """数字+班"""
        assert _extract_class_number("2510班") == 2510

    def test_chinese_prefix(self):
        """中文前缀+数字+班"""
        assert _extract_class_number("高三1班") == 1

    def test_no_digits(self):
        """无数字返回 0"""
        assert _extract_class_number("无数字") == 0

    def test_empty_string(self):
        """空字符串返回 0"""
        assert _extract_class_number("") == 0

    def test_multiple_digit_groups(self):
        """多个数字组拼接"""
        assert _extract_class_number("2024级1班") == 20241


# =============================================================================
# _filter_subjects
# =============================================================================


class TestFilterSubjects:
    """科目列表过滤"""

    def test_physics_direction(self):
        """物理方向不应包含历史"""
        result = _filter_subjects("物理")
        assert "历史" not in result
        assert "物理" in result

    def test_history_direction(self):
        """历史方向不应包含物理"""
        result = _filter_subjects("历史")
        assert "物理" not in result
        assert "历史" in result

    def test_unknown_direction(self):
        """未知方向返回全部科目"""
        result = _filter_subjects("")
        assert "物理" in result
        assert "历史" in result
        assert len(result) == 11  # 全部 11 科
